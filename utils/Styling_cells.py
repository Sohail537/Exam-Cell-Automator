from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font
def sgpa_styling(file):
    wb=load_workbook(file)
    for ws in wb.sheetnames:
        sheet = wb[ws]
        
        # Auto-size columns based on content        
        if "Analysis" in ws:
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    if isinstance(cell.value, int):
                        break
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass

                adjusted_width = (max_length + 2)   # Adjust as needed
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width
            if ws!="Overall Analysis":
                sheet.merge_cells("I1:L1")
                sheet["I1"]="Sem Toppers"
                sheet["I1"].font=Font(bold=True)
                sheet["I1"].alignment=Alignment(horizontal='center')
        else:
            for column in sheet.columns:
                max_length = 0                
                cell_object = column[0]  # Assuming you want to get the column name from the first row
                column_name = cell_object.value  # Assuming the first row contains the column names
                if column_name == "Roll No":
                    sheet.column_dimensions[column[0].column_letter].width = 11
                elif column_name=="Total Credits":
                    sheet.column_dimensions[column[0].column_letter].width = 12
                elif column_name=="Pass Percentage":
                    sheet.column_dimensions[column[0].column_letter].width = 15
                elif column_name=="GBM":
                    sheet.column_dimensions[column[0].column_letter].width = 5
                elif column_name=="Status" or column_name=="Backlogs" or column_name=="Points" or column_name=="SGPA":
                    pass
                else:
                    sheet.column_dimensions[column[0].column_letter].width = 15
                cell_object.alignment = Alignment(wrap_text=True)
                    
    wb.save(file)
def cgpa_styling(file):
    wb=load_workbook(file)
    for ws in wb.sheetnames:
        sheet = wb[ws]
        for column in sheet.columns:
            max_length = 0                
            cell_object = column[0]  # Assuming you want to get the column name from the first row
            column_name = cell_object.value  # Assuming the first row contains the column names
            if column_name == "Roll No":
                sheet.column_dimensions[column[0].column_letter].width = 11
            cell_object.alignment = Alignment(wrap_text=True)
    wb.save(file)