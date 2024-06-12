from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from utils.students_filter import require_htno
import openpyxl 
import math
from datetime import datetime






def create_pdf(filename,portal_data,seating_plan):
    workbook = openpyxl.Workbook()
    wb = openpyxl.load_workbook(seating_plan)
    sheet = workbook.active
    sheet2 = wb.active

    sheet2 = wb["Sheet1"]
    sheet1 = wb["Sheet2"]

    
    # Create a canvas object with A4 size in landscape orientation
    c = canvas.Canvas(filename, pagesize=landscape(A4))


    data = require_htno(portal_data)
    sheet2_row = sheet2.max_row

    for k in range(2, sheet2_row + 1):
        if sheet2.cell(k, 2).value == "Regular":
            sheet_name = str(sheet2.cell(k, 1).value)
            count = 0
            length = length_dataFrame(sheet2.cell(k, 1).value, sheet2.cell(k, 3).value,portal_data)
            x = length / 36
            y = math.ceil(x)

            for page_num in range(y):  # Loop over pages
                # Calculate positions of the rectangles

                #rectangle measures
                left_margin_cm = 0.8  # in cm
                top_margin_cm = 0.15  # in cm
                width_cm = 28  # in cm
                height_cm = 0.7 # in cm
                spacing_cm = 0.5  # in cm

                # Convert margins and dimensions from cm to points
                left_margin = left_margin_cm * 28.35
                top_margin = top_margin_cm * 28.35
                width = width_cm * 28.35
                height = height_cm * 28.35
                spacing = spacing_cm * 28.35

                page_width, page_height = landscape(A4)
                rect1_x = left_margin
                rect1_y = page_height - top_margin - height
                rect2_x = left_margin
                rect2_y = rect1_y - height 
                rect3_x = left_margin
                rect3_y = rect2_y - height
                rect4_x = left_margin
                rect4_y = rect3_y - height

                rect_height=rect1_y - height

                
                
                #Small box measures
                box_width=110
                box_padding=10

                # Set up some general styling
                c.setFont("Helvetica-Bold", 14)


                c.rect(rect1_x, rect1_y, width, height, fill=0)
                text1="JAWAHARLAL NEHRU TECHNOLOGICAL UNIVERSITY - KAKINADA - 533 003"
                c.drawString(rect1_x+100,page_height-top_margin-height+5,text1)


                # Draw the second rectangle
                c.rect(rect2_x, rect2_y, width-150, height, fill=0)
                text2="Name of the Examination Centre" 
                c.drawString(rect2_x+200,page_height-top_margin-(2*height)+5,text2)

                c.drawString(rect2_x+150,page_height-top_margin-(3*height)+5,sheet1.cell(2, 1).value)

                # Draw the side rectangle for second rectangle
                c.rect(width-150+left_margin,rect2_y - height,150,height,fill=0)
                text3="College Code"
                c.drawString(width-150+left_margin+20,page_height-top_margin-(2*height)+5,text3)

                c.drawString(width-150+left_margin+60,page_height-top_margin-(3*height)+5,str(sheet1.cell(2, 2).value))


                # Draw the third rectangle
                c.rect(rect3_x, rect3_y, width-150, height, fill=0)

                # Draw the side rectangle for third rectangle
                c.rect(width-150+left_margin,rect_height,150,height,fill=0)
                


                # Draw the fourth rectangle
                c.rect(rect4_x, rect4_y, width, height, fill=0)

                c.drawString(rect2_x+200,page_height-top_margin-(4*height)+5,sheet1.cell(2, 3).value)


                #Texts
                text4="SEATING ARRANGEMENT"
                c.drawString((page_width-200)/2,page_height-top_margin-(5*height)+6,text4)
                c.line((page_width-200)/2,496,500,496)
                

                text5="Regulation:"
                c.drawString(122,page_height-top_margin-(6*height)+16,text5)


                c.drawString(200,page_height-top_margin-(6*height)+16,sheet2.cell(k, 3).value[0:3])


                text6="HALL No:"
                c.drawString(550,page_height-top_margin-(6*height)+16,text6)

                parts = sheet2.cell(k, 6).value.split(',')
                result = parts[page_num]

                c.drawString(620,page_height-top_margin-(6*height)+16,result)


                text7="Date of Examination:"
                c.drawString(58,page_height-top_margin-(7*height)+16,text7)

                input_date_str = sheet1.cell(2, 4).value

                output_date_str = input_date_str.strftime("%d-%m-%Y")

                c.drawString(200,page_height-top_margin-(7*height)+16,str(output_date_str))


                text8="Session:"
                c.drawString(555,page_height-top_margin-(7*height)+16,text8)

                c.drawString(620,page_height-top_margin-(7*height)+16,str(sheet2.cell(k, 8).value))

                #Bold line 
                c.line(left_margin, 463, left_margin+width, 463)
                c.line(left_margin, 464, left_margin+width, 464)
                c.line(left_margin, 465, left_margin+width, 465)


                for i in range(0,6):
        
                    text9="H.T.No"
                    c.drawString(left_margin,page_height-(((9+i)+i)*height)+-(5*i),text9)

                    text10="Q.P.Set No."
                    c.drawString(left_margin,page_height-(((i+10)+i)*height)+-(5*i),text10)
                    
                    
                    for j in range(1,7):
                
                        c.rect((box_width*j)+(box_padding*(j-1)),page_height-top_margin-(((i+9)+i)*height)-(5*i),box_width,height)
                        c.rect((box_width*j)+(box_padding*(j-1)),page_height-top_margin-(((i+10)+i)*height)-(5*i),box_width,height)    
                
                #Bold line

                c.line(left_margin, 160, left_margin+width, 160)
                c.line(left_margin, 161, left_margin+width, 161)
                c.line(left_margin, 162, left_margin+width, 162)

                
                #Note

                text11="Note:Round off the HALLTICKET number of ABSENTEE candidateswith RED INK pen"
                c.drawString(100+left_margin,page_height-height+5-height-height-height-height-height-height-height-height-height-height-height-(11*height),text11)
                

                #total

                text12="TOTAL"
                c.drawString(left_margin+50,page_height-top_margin-(23*height)-20,text12)

                text13="Present:"
                c.drawString(left_margin+300,page_height-top_margin-(23*height)-20,text13)

                c.line(left_margin+357,111,left_margin+457,111)

                text14="Absent:"
                c.drawString(left_margin+550,page_height-top_margin-(23*height)-20,text14)

                c.line(left_margin+605,111,left_margin+705,111)


                #Invigilator

                text15="Name of the  1."
                c.drawString(left_margin+25,page_height-top_margin-(25*height)-20,text15)

                c.line(left_margin+125,75,left_margin+300,75)


                text16="Signature of the  1."
                c.drawString(left_margin+450,page_height-top_margin-(25*height)-20,text16)

                c.line(left_margin+575,75,left_margin+750,75)

                text17="Invigilator(s) 2."
                c.drawString(left_margin+25,page_height-top_margin-(26*height)-20,text17)

                c.line(left_margin+125,54,left_margin+300,54)

                text18="Invigilator(s)       2."
                c.drawString(left_margin+450,page_height-top_margin-(26*height)-20,text18)

                c.line(left_margin+575,54,left_margin+750,54)


                #OIE
                
                
                text19="OIE"
                c.drawString(left_margin+50,page_height-top_margin-(28*height)-20,text19)

                #CHIEF SUPERINTENDENT
                
                text20="CHIEF SUPERINTENDENT"
                c.drawString(left_margin+600,page_height-top_margin-(28*height)-23,text20)




                for i in range(0,6):
                    
                    text9="H.T.No"
                    c.drawString(left_margin,page_height-(((9+i)+i)*height)+-(5*i),text9)

                    text10="Q.P.Set No."
                    c.drawString(left_margin,page_height-(((i+10)+i)*height)+-(5*i),text10)
                    
                    
                    for j in range(1,7):
                
                        c.rect((box_width*j)+(box_padding*(j-1)),page_height-top_margin-(((i+9)+i)*height)-(5*i),box_width,height)
                        c.rect((box_width*j)+(box_padding*(j-1)),page_height-top_margin-(((i+10)+i)*height)-(5*i),box_width,height)    

                set = sheet2.cell(k, 5).value
                
                for i in range(0, 6):
                    for j in range(1, 7):
                        if count < len(data['0'+str(sheet2.cell(k, 1).value)][sheet2.cell(k, 3).value]):
                            c.drawString((box_width * (i+1)) + (box_padding * (i)) + 15,
                                         page_height - top_margin - (((j + 7) + j) * height) - (5 * j) + 10,
                                         data['0'+str(sheet2.cell(k, 1).value)][sheet2.cell(k, 3).value][count])
                            if sheet2.cell(k, 4).value == "Yes":
                                if set >= 5 :
                                    set = 1
                                    c.drawString((box_width * (i+1)) + (box_padding * (i)) + 50,
                                         page_height - top_margin - (((j + 8) + j) * height) - (5 * j) + 10,
                                         str(set))
                                    set = set + 1
                                else :
                                    c.drawString((box_width * (i+1)) + (box_padding * (i)) + 50,
                                         page_height - top_margin - (((j + 8) + j) * height) - (5 * j) + 10,
                                         str(set))
                                    set = set + 1
                            else :
                                c.drawString((box_width * (i+1)) + (box_padding * (i)) + 50,
                                         page_height - top_margin - (((j + 8) + j) * height) - (5 * j) + 10,
                                         str(set))

                            
                            count += 1

                c.showPage()  # Show the page and start a new one
                print("PDF created successfully.")

    # Save the PDF
    c.save()


def length_dataFrame(branch_code,subject_code,portal_data):
    data = require_htno(portal_data)
    length = len(data['0'+str(branch_code)][subject_code])
    return length
