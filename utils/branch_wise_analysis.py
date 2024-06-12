from pandas import *
import matplotlib 
matplotlib.use('agg')
import matplotlib.pyplot as plt
from openpyxl import load_workbook,drawing
def delete_branch (file,branch):
    wb = load_workbook(file)
    sheets=wb.sheetnames
    for i in sheets:
        if i != branch and i != branch + ' Analysis':
            wb.remove(wb[i])
    wb.save(file)

def branchwise_analysis(file,branch,grades):
    data=read_excel(file,sheet_name=[branch])
    data=DataFrame(data[branch])
    delete_branch(file,branch)
    for i in list(data.columns)[1:-7]:
        final_labels=[]
        final_grades=[]
        grades_of_subject=data[i].tolist()
        subject_grades_df=DataFrame({'Roll No':data['Roll No'],i:grades_of_subject})
        students_count=[]
        for j in grades:
            students_count.append(grades_of_subject.count(j[0]))
        for j in range(len(students_count)):
            if students_count[j]!=0:
                final_grades.append(students_count[j])
                final_labels.append(grades[j][0])
        subject_grades_count_df=DataFrame(columns=['Grades','No.of Students'])
        for j in range(len(students_count)):
            subject_grades_count_df.loc[len(subject_grades_count_df.index)]=[grades[j][0],students_count[j]]
        plt.pie(final_grades, labels=final_labels, autopct="%.2f%%")
        strFile="./Piechart.png"
        plt.title(i)
        plt.savefig(strFile)
        plt.clf()
        plt.close()
        subcode=i.split(" ")[-1]
        writer=ExcelWriter(file,engine='openpyxl',mode='a',if_sheet_exists="overlay")
        subject_grades_df.to_excel(writer,sheet_name=subcode,index=False)
        subject_grades_count_df.to_excel(writer,sheet_name=subcode,index=False,startcol=3,startrow=1)
        workbook =writer.book
        ws = workbook[subcode]
        img = drawing.image.Image(strFile)
        img.anchor = 'F2'
        ws.add_image(img)
        writer.close()