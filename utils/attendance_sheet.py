from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from utils.students_filter import require_htno
import openpyxl 
import math




def generate_attendance_sheet(filename,portal_data,seating_plan):

    workbook = openpyxl.Workbook()
    wb = openpyxl.load_workbook(seating_plan)
    sheet = workbook.active
    sheet2 = wb.active

    sheet2 = wb["Sheet1"]
    sheet1 = wb["Sheet2"]

    # Create a canvas object with A4 size in landscape orientation
    c = canvas.Canvas(filename, pagesize= A4)

    data = require_htno(portal_data)
    sheet2_row = sheet2.max_row

    for k in range(2, sheet2_row + 1):
        if sheet2.cell(k, 2).value == "Regular":
            sheet_name = str(sheet2.cell(k, 1).value)
            count = 0
            length = length_dataFrame(sheet2.cell(k, 1).value, sheet2.cell(k, 3).value,portal_data)
            x = length / 18
            y = math.ceil(x)
            # no = 1

            for page_num in range(y):

                if page_num % 2 == 0:  # Reset numbering every 2 pages
                    no = 1
    
                #rectangle measures
                left_margin_cm = 0.8  # in cm
                top_margin_cm = 0.15  # in cm
                width_cm = 28  # in cm
                height_cm = 0.8 # in cm
                spacing_cm = 0.5  # in cm

                # Convert margins and dimensions from cm to points
                left_margin = left_margin_cm * 28.35
                top_margin = top_margin_cm * 28.35
                width = width_cm * 28.35
                height = height_cm * 28.35
                spacing = spacing_cm * 28.35

                # Calculate positions of the rectangles
                page_width, page_height = A4
                print(page_width, page_height)
                rect1_x = left_margin
                rect1_y = page_height - top_margin - height
                rect2_x = left_margin
                rect2_y = rect1_y - height 
                rect3_x = left_margin
                rect3_y = rect2_y - height
                rect4_x = left_margin
                rect4_y = rect3_y - height

                
                rect_height=rect1_y - height
                rect_width=width-(width-150)

                

                # Set up some general styling
                c.setFont("Helvetica-Bold", 14)

                # Draw the first rectangle
                c.rect(rect1_x, rect1_y, page_width-(2*left_margin), height, fill=0)
                text1="JAWAHARLAL NEHRU TECHNOLOGICAL UNIVERSITY - KAKINADA - 533 003"
                c.drawString(45,page_height-top_margin-height+5,text1)


                # Draw the second rectangle
                c.rect(rect2_x, rect2_y, page_width-(2*left_margin)-100, height, fill=0)
                text2="Name of the Examination Centre"
                c.drawString(rect2_x+150,page_height-top_margin-(2*height)+5,text2)

                c.drawString(rect2_x+100,page_height-top_margin-(3*height)+5,sheet1.cell(2, 1).value)

                # Draw the side rectangle for second rectangle
                c.rect(page_width-(2*left_margin)-100+left_margin,rect2_y,100,height,fill=0)
                text3="College Code"
                c.drawString(page_width-(2*left_margin)-100+left_margin+7,page_height-top_margin-(2*height)+5,text3)

                c.drawString(page_width-(2*left_margin)-25,page_height-top_margin-(3*height)+5,str(sheet1.cell(2, 2).value))


                # Draw the third rectangle
                c.rect(rect3_x, rect3_y, page_width-(2*left_margin)-100, height, fill=0)

                # Draw the side rectangle for third rectangle
                c.rect(page_width-(2*left_margin)-100+left_margin,rect3_y,100,height,fill=0)
                
                c.setFont("Helvetica-Bold", 11)

                #Texts
                text4="HALL-WISE ATTENDANCE OF CANDIDATES AND INFORMATION RELATING TO ANSWER BOOKS"
                c.drawString(left_margin+20,page_height-top_margin-(4*height),text4)
                # c.line((page_width-200)/2,496,500,496)

                c.setFont("Helvetica-Bold", 14)    

                text5="Name of Exam:"
                c.drawString(70,page_height-top_margin-(6*height)+16,text5)

                c.drawString(170,page_height-top_margin-(6*height)+16,sheet1.cell(2, 3).value)


                text6=" Date:"
                c.drawString(400,page_height-top_margin-(6*height)+16,text6)

                input_date_str = sheet1.cell(2, 4).value

                output_date_str = input_date_str.strftime("%d-%m-%Y")

                c.drawString(440,page_height-top_margin-(6*height)+16,str(output_date_str))


                text7="Subject:"
                c.drawString(116,page_height-top_margin-(7*height)+16,text7)
                # data1 = pdfToDataframe(r'c:\Users\HP\Downloads\03_1233931.pdf')

                # for index, row in data1.iterrows():
                #     if row['subcode'] == sheet2.cell(k, 3).value:
                        
                #         subject_name = row['subname']
                #         break
                

                c.drawString(170,page_height-top_margin-(7*height)+16,sheet2.cell(k, 9).value)


                text8="Lecture Hall:"
                c.drawString(355,page_height-top_margin-(7*height)+16,text8)

                parts = sheet2.cell(k, 6).value.split(',')
                # result = parts[page_num]
                # print(len(parts))


                if page_num % y < 2:
                    lecture_hall_text = parts[0]
                else:
                    lecture_hall_text = parts[1]
        
                c.drawString(440,page_height-top_margin-(7*height)+16, lecture_hall_text)


                c.rect(left_margin, page_height-top_margin-(9*height),page_width-(2*left_margin), (2*height)+10)

                set = sheet2.cell(k, 5).value                

                for i in range(10,28):
                    if count < len(data['0'+str(sheet2.cell(k, 1).value)][sheet2.cell(k, 3).value]):
                        c.rect((left_margin), page_height-top_margin-(i*height),page_width-(2*left_margin), height)
                        c.drawString((4*left_margin),page_height-top_margin-((i+1)*height)+30,data['0'+str(sheet2.cell(k, 1).value)][sheet2.cell(k, 3).value][count])
                        count += 1
                    
                        c.drawString((left_margin)+10,page_height-top_margin-((i+1)*height)+30,str(no))
                        no = no +1
                        if sheet2.cell(k, 4).value == "Yes":
                            if set >= 5 :
                                    set = 1
                                    c.drawString((13*left_margin)+50,page_height-top_margin-((i+1)*height)+30,str(set))
                                    set = set + 1
                            else :
                                    c.drawString((13*left_margin)+50,page_height-top_margin-((i+1)*height)+30,str(set))
                                    set = set + 1
                        else :
                            c.drawString((13*left_margin)+50,page_height-top_margin-((i+1)*height)+30,str(set))

                        
                        c.line((left_margin)+50,page_height-top_margin-(7*height)+10,(left_margin)+50,page_height-top_margin-(i*height))

                        c.line((7*left_margin)+50,page_height-top_margin-(7*height)+10,(7*left_margin)+50,page_height-top_margin-(i*height))

                        c.line((11*left_margin)+55,page_height-top_margin-(7*height)+10,(11*left_margin)+55,page_height-top_margin-(i*height))

                        c.line((15*left_margin)+50,page_height-top_margin-(7*height)+10,(15*left_margin)+50,page_height-top_margin-(i*height))

                        size = i
                
                text="No. Allotted:"
                c.drawString(left_margin,page_height-top_margin-((size+3)*height)+20,text)

                text="No. Absent:"
                c.drawString(10*left_margin,page_height-top_margin-((size+3)*height)+20,text)

                text="No. Present:"
                c.drawString(20*left_margin,page_height-top_margin-((size+3)*height)+20,text)

                text="Note: Absentees are rounded in RED ink"
                c.drawString(left_margin,page_height-top_margin-((size+5)*height)+20,text)

                c.line(left_margin,page_height-((size+4)*height)-10,(left_margin)+28,page_height-((size+4)*height)-10)

                text="Name & Signature of Invigilator"
                c.drawString(left_margin,page_height-top_margin-((size+8)*height)+20,text)

                text=" Signature of Chief Superintendent"
                c.drawString(15*left_margin,page_height-top_margin-((size+8)*height)+20,text)


                c.drawString((left_margin)+10,page_height-top_margin-(10*height)+35,"S.No")

                c.drawString((4*left_margin),page_height-top_margin-(10*height)+35,"Hall Ticket No")

                c.drawString((9*left_margin)+10,page_height-top_margin-(10*height)+45,"Main Answer")

                c.drawString((9*left_margin)+10,page_height-top_margin-(10*height)+30,"Book S.No.")

                c.drawString((13*left_margin)+16,page_height-top_margin-(10*height)+60,"Question")

                c.drawString((13*left_margin)+16,page_height-top_margin-(10*height)+45,"Paper Set")

                c.drawString((14*left_margin)+15,page_height-top_margin-(10*height)+30,"No.")

                c.drawString(18*left_margin,page_height-top_margin-(10*height)+35,"Signature of Candidate")

                

    


                c.showPage()  # Show the page and start a new one
                print("PDF created successfully.")

    c.save()


def length_dataFrame(branch_code,subject_code,portal_data):
    data = require_htno(portal_data)
    length = len(data['0'+str(branch_code)][subject_code])
    return length
