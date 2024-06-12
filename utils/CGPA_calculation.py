from pandas import *
from openpyxl import load_workbook
from utils.CGPA_class import CGPA
objects_list=[]
def cgpa_calculation(input_excel,sem):
    global objects_list
    input_data=load_workbook(input_excel)
    branches_in_input_data=input_data.sheetnames
    print(branches_in_input_data)
    for branch in branches_in_input_data:
        if "Analysis" not in branch and branches_in_input_data[-1]=="Overall Analysis":
            branch_wise_data=read_excel(input_excel,sheet_name=branch)
            if len(objects_list)==0:
                objects_list.append(CGPA(branch_wise_data.iloc[0,0][6:8]))
            for obj in objects_list:
                if obj.branch ==  branch_wise_data.iloc[0,0][6:8]:
                    print(obj.branch)
                    branch_wise_data=branch_wise_data[["Roll No","Points","Total Credits","SGPA","Backlogs"]]
                    branch_wise_data=branch_wise_data.rename(columns={"Total Credits":"Total Credits sem"+str(sem),"Points":"Points sem"+str(sem),"SGPA":"SGPA sem"+str(sem),"Backlogs":"Backlogs sem"+str(sem)})
                    if len(obj.cgpa_df.columns)==0:
                        obj.cgpa_df=branch_wise_data
                    else:
                        obj.cgpa_df=merge(obj.cgpa_df,branch_wise_data,"outer",on="Roll No")
                    obj.cgpa_df=obj.cgpa_df.fillna(0)
                    break
            else:
                objects_list.append(CGPA(branch_wise_data.iloc[0,0][6:8]))
                obj=objects_list[-1]
                print(obj.branch)
                branch_wise_data=branch_wise_data[["Roll No","Points","Total Credits","SGPA","Backlogs"]]
                branch_wise_data=branch_wise_data.rename(columns={"Total Credits":"Total Credits sem"+str(sem),"Points":"Points sem"+str(sem),"SGPA":"SGPA sem"+str(sem),"Backlogs":"Backlogs sem"+str(sem)})
                if len(obj.cgpa_df.columns)==0:
                    obj.cgpa_df=branch_wise_data
                else:
                    obj.cgpa_df=merge(obj.cgpa_df,branch_wise_data,"outer",on="Roll No")
                obj.cgpa_df=obj.cgpa_df.fillna(0)
        elif branches_in_input_data[-1]!="Overall Analysis":
            branch_wise_data=read_excel(input_excel,sheet_name=branches_in_input_data[0])
            if len(objects_list)==0:
                objects_list.append(CGPA(branch_wise_data.iloc[0,0][6:8]))
            for obj in objects_list:
                if obj.branch ==  branch_wise_data.iloc[0,0][6:8]:
                    print(obj.branch)
                    branch_wise_data=branch_wise_data[["Roll No","Points","Total Credits","SGPA","Backlogs"]]
                    branch_wise_data=branch_wise_data.rename(columns={"Total Credits":"Total Credits sem"+str(sem),"Points":"Points sem"+str(sem),"SGPA":"SGPA sem"+str(sem),"Backlogs":"Backlogs sem"+str(sem)})
                    if len(obj.cgpa_df.columns)==0:
                        obj.cgpa_df=branch_wise_data
                    else:
                        obj.cgpa_df=merge(obj.cgpa_df,branch_wise_data,"outer",on="Roll No")
                    obj.cgpa_df=obj.cgpa_df.fillna(0)
                    break
            break
